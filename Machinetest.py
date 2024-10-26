import openpyxl
import os

file_name = "user_data.xlsx"

# Check if the Excel file exists; if not, create it with headers
if not os.path.exists(file_name):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Name", "Email", "Phone Number"])
    workbook.save(file_name)

#  add a user
def add_user():
    name = input("Enter Name: ")
    email = input("Enter Email: ")
    phone = input("Enter Phone Number: ")
    
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    sheet.append([name, email, phone])
    workbook.save(file_name)
    print("User added successfully!\n")

#  display all users
def display_users():
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    print("\nStored Users:")
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(f"Name: {row[0]}, Email: {row[1]}, Phone: {row[2]}")
    print("")

# Main menu
while True:
    print("Choose an option:")
    print("1. Add User")
    print("2. Display Users")
    print("3. Exit")
    choice = input("Enter your choice (1/2/3): ")
    
    if choice == '1':
        add_user()
    elif choice == '2':
        display_users()
    elif choice == '3':
        print("Exiting program.")
        break
    else:
        print("Invalid choice, please try again.\n")
